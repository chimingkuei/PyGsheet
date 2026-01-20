import gspread
import logging
from pathlib import Path
from datetime import datetime, timedelta
import schedule
from openpyxl import Workbook, load_workbook
from __init__ import __version__

# =====================
# 基本路徑
# =====================
BASE_DIR = Path(__file__).resolve().parent
LOG_DIR = BASE_DIR / "Logger"
EXPORT_DIR = BASE_DIR / "Export"
LOG_DIR.mkdir(exist_ok=True)
EXPORT_DIR.mkdir(exist_ok=True)

# =====================
# Logger
# =====================
def setup_logger():
    today = datetime.now().strftime("%Y-%m-%d")
    log_file = LOG_DIR / f"Log_{today}.txt"

    logger = logging.getLogger("daily_logger")
    logger.setLevel(logging.INFO)
    if logger.handlers:
        return logger

    formatter = logging.Formatter("%(asctime)s | %(levelname)s | %(message)s")
    fh = logging.FileHandler(log_file, encoding="utf-8")
    sh = logging.StreamHandler()
    fh.setFormatter(formatter)
    sh.setFormatter(formatter)
    logger.addHandler(fh)
    logger.addHandler(sh)
    return logger

logger = setup_logger()
logger.info(f"程式啟動 | 版本：{__version__}")

# =====================
# 每日 Excel
# =====================
def get_daily_workbook():
    today_str = datetime.now().strftime("%Y-%m-%d")
    file_path = EXPORT_DIR / f"Summary_{today_str}.xlsx"

    if file_path.exists():
        wb = load_workbook(file_path)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Summary"
        # 新增 J 欄
        ws.append(["日期", "組別", "成員", "起始列", "結束列", "午餐每日總和", "飲料每日總和"])
        wb.save(file_path)
    return wb, ws, file_path

# =====================
# 判斷月底（測試用）
# =====================
def is_month_end():
    # 真實環境可用下面註解：
    # today = datetime.now()
    # tomorrow = today + timedelta(days=1)
    # return tomorrow.month != today.month and today.weekday() < 5
    return True  # 測試用：每次都算月底

# =====================
# 月底彙總（依 GROUP 排序）
# =====================
def generate_monthly_summary():
    now = datetime.now()
    month_str = now.strftime("%Y-%m")
    monthly_file = EXPORT_DIR / f"Monthly_Summary_{month_str}.xlsx"

    wb_month = Workbook()
    ws_month = wb_month.active
    ws_month.title = "Monthly Summary"

    # 標題
    ws_month.append(["組別", "成員", "午餐月累計", "飲料月累計"])

    # key = (member, group_no)
    cumulative = {}

    # 讀取本月所有每日檔
    for file in sorted(EXPORT_DIR.glob(f"Summary_{month_str}-*.xlsx")):
        wb = load_workbook(file)
        ws = wb.active

        for row in ws.iter_rows(min_row=2, values_only=True):
            # [日期, 組別, 成員, 起始列, 結束列, E每日總和, J每日總和]
            _, group_no, member, _, _, daily_e, daily_j = row
            key = (member, group_no)
            if key not in cumulative:
                cumulative[key] = {"E": 0, "J": 0}
            cumulative[key]["E"] += daily_e or 0
            cumulative[key]["J"] += daily_j or 0

    # ⭐ 依 GROUP 排序（主），成員排序（次）
    for (member, group_no), totals in sorted(
        cumulative.items(),
        key=lambda x: (x[0][1], x[0][0])  # (group_no, member)
    ):
        ws_month.append([group_no, member, totals["E"], totals["J"]])

    wb_month.save(monthly_file)
    logger.info(f"月底彙總完成（依 GROUP 排序）：{monthly_file.name}")

# =====================
# 取得今日 Sheet 名稱 (例如 01.20(二))
# =====================
def get_today_gsheet_sheet(gc, share_link):
    sh = gc.open_by_url(share_link)
    today_str = datetime.now().strftime("%m.%d")  # 例如 "01.20"
    # 模糊匹配 sheet 名稱
    sheet_name = None
    for name in sh.worksheets():
        if today_str in name.title:
            sheet_name = name
            break
    if sheet_name is None:
        raise ValueError(f"找不到今天的 Sheet ({today_str})")
    return sheet_name

# =====================
# 主工作
# =====================
def job():
    today_weekday = datetime.now().weekday()
    if today_weekday >= 5:
        logger.info("今天是週六 / 週日，job 不執行")
        return

    logger.info("job 開始")
    try:
        share_link = "https://docs.google.com/spreadsheets/d/1k2yjKUA4m886_Vj24V_jnpt4ClTkQIjoabBBQKb4Jko/edit"
        gc = gspread.service_account(filename="token.json")

        worksheet = get_today_gsheet_sheet(gc, share_link)

        START_ROW = 5
        GROUP_SIZE = 4
        GROUP_COUNT = 18
        TOTAL_COUNT = GROUP_SIZE * GROUP_COUNT

        # E 欄
        raw_e = worksheet.col_values(5)[START_ROW - 1 : START_ROW - 1 + TOTAL_COUNT]
        numbers_e = []
        for v in raw_e:
            try:
                numbers_e.append(float(v))
            except:
                numbers_e.append(0)

        # J 欄
        raw_j = worksheet.col_values(10)[START_ROW - 1 : START_ROW - 1 + TOTAL_COUNT]
        numbers_j = []
        for v in raw_j:
            try:
                numbers_j.append(float(v))
            except:
                numbers_j.append(0)

        # 補齊長度
        while len(numbers_e) < TOTAL_COUNT:
            numbers_e.append(0)
        while len(numbers_j) < TOTAL_COUNT:
            numbers_j.append(0)

        # B 欄成員
        member_values = worksheet.col_values(2)

        wb, ws, file_path = get_daily_workbook()
        today_str = datetime.now().strftime("%Y-%m-%d")

        for group_index in range(GROUP_COUNT):
            start_idx = group_index * GROUP_SIZE

            group_e = numbers_e[start_idx : start_idx + GROUP_SIZE]
            group_j = numbers_j[start_idx : start_idx + GROUP_SIZE]

            group_sum_e = sum(group_e)
            group_sum_j = sum(group_j)

            start_row = START_ROW + start_idx
            end_row = start_row + GROUP_SIZE - 1
            member_name = (
                member_values[start_row - 1]
                if start_row - 1 < len(member_values)
                else ""
            )

            ws.append([
                today_str,
                group_index + 1,
                member_name,
                start_row,
                end_row,
                group_sum_e,
                group_sum_j
            ])

            logger.info(
                f"第 {group_index+1:02d} 組 | 成員={member_name} | "
                f"E{start_row}~E{end_row}={group_sum_e} | "
                f"J{start_row}~J{end_row}={group_sum_j}"
            )

        wb.save(file_path)
        logger.info(f"日報完成：{file_path.name}")

        # 月底彙總
        if is_month_end():
            logger.info("模擬月底，開始生成月結")
            generate_monthly_summary()

        logger.info("job 結束")

    except Exception:
        logger.exception("job 發生錯誤")

# =====================
# schedule
# =====================
schedule.every().day.at("11:29").do(job)

# =====================
# 主迴圈
# =====================
if __name__ == "__main__":
    import time
    while True:
        schedule.run_pending()
        time.sleep(30)
